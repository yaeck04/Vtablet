import flet as ft
import json
import os
import random
import calendar
import io
import base64
import threading
import time 
import requests
from datetime import datetime, timedelta
from PIL import Image, ImageDraw, ImageFont
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import urllib3 

# Desactivar advertencias de inseguridad SSL/Urllib3 si es necesario
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Constantes
ARCHIVO_JSON = "sorteos_unificados_con_fijos.json"
ARCHIVO_REGLAS = "reglas_color.json"

# PALETA DE COLORES
PALETA_COLORES = [
    "#FF0000", "#00FF00", "#0000FF", "#FFFF00", "#00FFFF", "#FF00FF",
    "#FFA500", "#800080", "#A52A2A", "#000000", "#808080", "#FFFFFF",
    "#4CAF50", "#2196F3", "#FF9800", "#9C27B0", "#F44336", "#E91E63",
    "#795548", "#607D8B", "#8BC34A", "#03A9F4", "#FFC107", "#9E9E9E"
]

# CONFIGURACIÓN ACTUALIZACIÓN DE DB
EMAIL = "khloealba932@gmail.com"
PASSWORD = "Anabelyae04"

URLS_TXT = {
    ("GA", "M"): "https://www.lotterycorner.com/results/download/ga-cash-3-midday-2026.txt",
    ("GA", "E"): "https://www.lotterycorner.com/results/download/ga-cash-3-evening-2026.txt",
    ("GA", "N"): "https://www.lotterycorner.com/results/download/ga-cash-3-night-2026.txt",
    ("FL", "M"): "https://www.lotterycorner.com/results/download/fl-pick-3-midday-2026.txt",
    ("FL", "E"): "https://www.lotterycorner.com/results/download/fl-pick-3-evening-2026.txt",
    ("NY", "M"): "https://www.lotterycorner.com/results/download/ny-numbers-midday-2026.txt",
    ("NY", "E"): "https://www.lotterycorner.com/results/download/ny-numbers-evening-2026.txt",
}

PRIORIDAD = {
    "GA_M": 0, "FL_M": 1, "NY_M": 2,
    "GA_E": 3, "FL_E": 4, "NY_E": 5,
    "GA_N": 6
}

TURNOS_POR_STATE = {
    "FL": ["M", "E"],
    "GA": ["M", "E", "N"],
    "NY": ["M", "E"]
}

COL_DIA_WIDTH = 50
COL_TURNO_WIDTH = 60

class ReglaColor:
    def __init__(self, tipo, digito, color):
        self.tipo = tipo 
        self.digito = str(digito)
        self.color = color

class LoteriaApp:
    def __init__(self, page: ft.Page):
        self.page = page
        self.page.title = "Lotería App"
        self.page.vertical_alignment = ft.MainAxisAlignment.START
        self.page.horizontal_alignment = ft.CrossAxisAlignment.CENTER
        self.page.scroll = ft.ScrollMode.AUTO 
        self.page.theme_mode = ft.ThemeMode.LIGHT
        self.page.padding = 10
        
        self.datos = self.cargar_datos()
        
        self.reglas_colores = []
        self.datos_filtrados = []
        self.anio_seleccionado = None
        self.mes_seleccionado = None 
        self.state_seleccionado = None
        self.color_seleccionado_temp = None 
        
        self.visualizacion_mode = 0 
        
        self.dd_anio = None
        self.dd_mes = None
        self.dd_state = None
        self.dd_visualizacion = None
        self.dd_regla_tipo = None
        self.input_digito_container = None 
        self.lista_reglas_ui = None
        self.contenedor_vista_tabla = None
        
        self.dialog_update = None
        self.progress_bar = None
        self.log_list_view = None 
        
        self.cargar_reglas_persistentes()
        self.configurar_ui()
        self.actualizar_lista_reglas()
    
    # Función Helper para agregar logs visuales
    def log(self, mensaje):
        self.log_list_view.controls.append(ft.Text(mensaje, size=12, color=ft.Colors.BLACK87))
        self.page.update()
    
    def generar_datos_dummy(self):
        dummy_data = []
        estados = ["FL", "GA", "NY"]
        fecha_base = datetime(2026,1,1)
        fin_anio = datetime(2026, 12, 31)
        
        fecha_actual = fecha_base
        while fecha_actual <= fin_anio:
            for estado in estados:
                turnos = TURNOS_POR_STATE[estado]
                for turno in turnos:
                    nums = sorted([random.randint(0, 9) for _ in range(3)])
                    n1, n2, n3 = nums
                    fijo1 = f"{n1}{n2}"
                    fijo2 = f"{n2}{n3}" 
                    
                    dummy_data.append({
                        "date": fecha_actual.strftime("%d/%m/%y"),
                        "state": estado,
                        "draw": turno,
                        "numbers": f"{n1}-{n2}-{n3}",
                        "fijos": [fijo1, fijo2]
                    })
            fecha_actual += timedelta(days=1)
        
        with open(ARCHIVO_JSON, "w", encoding="utf-8") as f:
            json.dump(dummy_data, f, indent=4)

    def cargar_datos(self):
        if not os.path.exists(ARCHIVO_JSON):
            self.generar_datos_dummy()
        with open(ARCHIVO_JSON, "r", encoding="utf-8") as f:
            return json.load(f)
    
    def cargar_reglas_persistentes(self):
        if os.path.exists(ARCHIVO_REGLAS):
            try:
                with open(ARCHIVO_REGLAS, "r", encoding="utf-8") as f:
                    reglas_data = json.load(f)
                    for r in reglas_data:
                        self.reglas_colores.append(ReglaColor(r["tipo"], r["digito"], r["color"]))
                print(f"Reglas cargadas: {len(self.reglas_colores)}")
            except Exception as e:
                print(f"Error cargando reglas: {e}")

    def guardar_reglas_persistentes(self):
        try:
            reglas_data = []
            for r in self.reglas_colores:
                reglas_data.append({
                    "tipo": r.tipo,
                    "digito": r.digito,
                    "color": r.color
                })
            with open(ARCHIVO_REGLAS, "w", encoding="utf-8") as f:
                json.dump(reglas_data, f, indent=4)
            print(f"Reglas guardadas en {ARCHIVO_REGLAS}")
        except Exception as e:
            print(f"Error guardando reglas: {e}")

    def configurar_ui(self):
        self.page.appbar = ft.AppBar(
            title=ft.Text("Lotería Visualizador", size=20, weight=ft.FontWeight.BOLD, color=ft.Colors.WHITE),
            center_title=True,
            bgcolor=ft.Colors.BLUE_800,
            actions=[ft.IconButton(ft.Icons.INFO, on_click=self.mostrar_ayuda, icon_color=ft.Colors.WHITE)]
        )
        
        # Crear componentes del diálogo de actualización
        self.log_list_view = ft.ListView(expand=True, height=150, spacing=2, auto_scroll=True)
        self.progress_bar = ft.ProgressBar(width=400)
        
        # Definir el diálogo
        self.dialog_update = ft.AlertDialog(
            modal=True,
            title=ft.Text("Actualizando Base de Datos"),
            content=ft.Column(
                [
                    ft.Text("Progreso:", size=12, weight=ft.FontWeight.BOLD),
                    self.progress_bar,
                    ft.Divider(height=10),
                    ft.Text("Detalles:", size=12, weight=ft.FontWeight.BOLD),
                    self.log_list_view,
                ],
                tight=True,
                height=300, 
                width=450,
                alignment=ft.MainAxisAlignment.START
            ),
            # SE AGREGA EL BOTÓN CERRAR MANUAL
            actions=[
                ft.TextButton("Cerrar", on_click=lambda _: self.close_dialog(self.dialog_update))
            ]
        )

        self.dd_anio = ft.Dropdown(label="Año", width=400, options=self.obtener_anios_opciones())
        opciones_mes = [ft.dropdown.Option("Todos", "Todos")] + [ft.dropdown.Option(str(i), calendar.month_name[i]) for i in range(1,13)]
        self.dd_mes = ft.Dropdown(label="Mes", width=400, options=opciones_mes, value="Todos")
        self.dd_state = ft.Dropdown(
            label="Lotería",
            width=400,
            options=[
                ft.dropdown.Option("FL", "Florida"),
                ft.dropdown.Option("GA", "Georgia"),
                ft.dropdown.Option("NY", "New York")
            ]
        )
        
        self.dd_visualizacion = ft.Dropdown(
            label="Vista",
            width=400,
            options=[
                ft.dropdown.Option("0", "Cubabingo"),
                ft.dropdown.Option("1", "Por Horario")
            ],
            value="0",
            on_select=self.cambiar_visualizacion
        )

        self.dd_regla_tipo = ft.Dropdown(
            label="Tipo Regla",
            width=400,
            options=[
                ft.dropdown.Option("completo", "Completo (Fijo exacto)"),
                ft.dropdown.Option("pareja", "Pareja (Ej: 22, 33)"),
                ft.dropdown.Option("digito", "Dígito (Decena o Terminal)"), 
                ft.dropdown.Option("decena", "Decena (1er dígito)"),
                ft.dropdown.Option("terminal", "Terminal (2º dígito)")
            ],
            on_select=self.actualizar_input_regla
        )
        
        self.input_digito_container = ft.Container(
            width=400,
            content=ft.Dropdown(label="Dígito", options=[ft.dropdown.Option(str(i), str(i)) for i in range(10)])
        )

        self.lista_reglas_ui = ft.ListView(expand=True, spacing=5, item_extent=50)
        
        self.contenedor_vista_tabla = ft.Column(
            scroll=ft.ScrollMode.AUTO, 
            horizontal_alignment=ft.CrossAxisAlignment.CENTER,
            alignment=ft.MainAxisAlignment.START
        )

        self.page.add(
            ft.Column(
                horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                controls=[
                    ft.Container(
                        padding=10,
                        bgcolor=ft.Colors.WHITE,
                        border_radius=10,
                        shadow=ft.BoxShadow(blur_radius=5, color=ft.Colors.GREY_300),
                        width=600, 
                        content=ft.Column(controls=[
                            ft.Text("Filtros y Vista", size=16, weight=ft.FontWeight.BOLD),
                            ft.Column(
                                controls=[self.dd_anio, self.dd_mes, self.dd_state, self.dd_visualizacion],
                                horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                                spacing=10
                            ),
                            ft.Row(
                                controls=[
                                    ft.Button("Generar Tabla", on_click=self.aplicar_filtros, expand=True, style=ft.ButtonStyle(bgcolor=ft.Colors.BLUE_600, color=ft.Colors.WHITE)),
                                    ft.Container(width=10),
                                    # Botón Actualizar DB
                                    ft.Button("Actualizar DB", icon=ft.Icons.REFRESH, on_click=self.iniciar_actualizacion, style=ft.ButtonStyle(bgcolor=ft.Colors.ORANGE_600, color=ft.Colors.WHITE))
                                ]
                            )
                        ])
                    ),
                    ft.Container(height=10),
                    
                    ft.Container(
                        padding=10,
                        bgcolor=ft.Colors.WHITE,
                        border_radius=10,
                        shadow=ft.BoxShadow(blur_radius=5, color=ft.Colors.GREY_300),
                        width=600,
                        content=ft.Column(controls=[
                            ft.Text("Reglas de Color", size=16, weight=ft.FontWeight.BOLD),
                            self.dd_regla_tipo,
                            self.input_digito_container,
                            ft.Row(controls=[
                                ft.Button("Elegir Color", on_click=self.mostrar_paleta_colores, expand=True, bgcolor=ft.Colors.GREY_200),
                                ft.Button("Agregar", on_click=self.agregar_regla, expand=True, bgcolor=ft.Colors.GREEN_600, color=ft.Colors.WHITE)
                            ]),
                            ft.Container(height=10),
                            ft.Text("Reglas Activas:", size=12, color=ft.Colors.GREY_600),
                            self.lista_reglas_ui
                        ])
                    ),
                    ft.Container(height=10),
                    
                    ft.Row(
                        wrap=True,
                        alignment=ft.MainAxisAlignment.CENTER,
                        controls=[
                            ft.Button("Excel", icon=ft.Icons.TABLE_CHART, on_click=self.exportar_excel, style=ft.ButtonStyle(bgcolor=ft.Colors.GREEN_700, color=ft.Colors.WHITE)),
                            ft.Container(width=10),
                            ft.Button("Imagen", icon=ft.Icons.IMAGE, on_click=self.exportar_imagen, style=ft.ButtonStyle(bgcolor=ft.Colors.PURPLE_700, color=ft.Colors.WHITE))
                        ]
                    ),
                    ft.Container(height=10),
                    
                    ft.Container(
                        padding=5,
                        bgcolor=ft.Colors.WHITE,
                        border_radius=5,
                        content=self.contenedor_vista_tabla
                    )
                ]
            )
        )
    
    # --- LÓGICA DE ACTUALIZACIÓN DE DB ---
    def iniciar_actualizacion(self, e):
        # Limpiar logs anteriores
        self.log_list_view.controls.clear()
        self.progress_bar.value = 0
        self.log("Iniciando proceso...")
        self.page.update()
        
        # Mostrar dialogo (Si ya está abierto, se reutiliza o muestra encima)
        self.page.show_dialog(self.dialog_update)
        
        # Ejecutar lógica pesada en hilo secundario
        self.page.run_thread(self.logica_actualizar_db)

    def logica_actualizar_db(self):
        try:
            self.log("🔐 Intentando conectar a Lotterycorner...")
            
            # Usar User-Agent para evitar bloqueos
            session = requests.Session()
            session.headers.update({
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
            })
            
            login_url = "https://www.lotterycorner.com/insider/login"
            login_data = {"email": EMAIL, "pwd": PASSWORD}
            
            try:
                r = session.post(login_url, data=login_data, timeout=15)
            except requests.exceptions.Timeout:
                raise Exception("❌ Timeout: No hubo respuesta en 15s. Revisa tu internet.")
            except requests.exceptions.ConnectionError:
                raise Exception("❌ Error de conexión: No se pudo llegar al servidor.")
            except Exception as e:
                raise Exception(f"❌ Error de red: {str(e)}")
            
            if "Logout" not in r.text:
                raise Exception("❌ Login fallido. Verifica usuario/contraseña o si hay captcha.")
            
            # Login OK
            self.progress_bar.value = 0.1
            self.log("✅ Login exitoso. Descargando archivos...")
            self.page.update()
            
            contenidos = {}
            pasos = len(URLS_TXT)
            i = 0
            
            for (estado, sorteo), url in URLS_TXT.items():
                i += 1
                self.log(f"⬇️ Descargando {estado}_{sorteo} ({i}/{pasos})...")
                # Actualizar barra progreso proporcional
                progreso = 0.1 + (i / pasos) * 0.6
                self.progress_bar.value = progreso
                self.page.update()
                
                try:
                    r = session.get(url, timeout=30)
                    r.raise_for_status()
                    contenidos[(estado, sorteo)] = r.text
                    self.log(f"   -> Descarga completada ({len(r.text)} bytes)")
                except Exception as ex:
                    self.log(f"   -> Error descargando {estado}_{sorteo}: {str(ex)}")
                
            # Parsear y Guardar
            self.log("🔨 Procesando datos y guardando...")
            self.progress_bar.value = 0.8
            self.page.update()
            
            combinaciones = {}
            if os.path.exists(ARCHIVO_JSON):
                with open(ARCHIVO_JSON, "r", encoding="utf-8") as f:
                    existentes = json.load(f)
                    for s in existentes:
                        clave = f"{s['date']}-{s['state']}-{s['draw']}"
                        combinaciones[clave] = s
            
            nuevos_contados = 0
            reemplazos = 0
            
            for (estado, sorteo), texto in contenidos.items():
                nuevos = self.parsear_txt(texto, estado, sorteo)
                for s in nuevos:
                    clave = f"{s['date']}-{s['state']}-{s['draw']}"
                    if clave in combinaciones:
                        if combinaciones[clave]["numbers"] != s["numbers"]:
                            combinaciones[clave] = s
                            reemplazos += 1
                    else:
                        combinaciones[clave] = s
                        nuevos_contados += 1

            todos = list(combinaciones.values())
            todos.sort(key=lambda s: (
                datetime.strptime(s["date"], "%d/%m/%y"),
                PRIORIDAD.get(f'{s["state"]}_{s["draw"]}', 99)
            ))

            with open(ARCHIVO_JSON, "w", encoding="utf-8") as f:
                json.dump(todos, f, indent=2)

            # Recargar en memoria
            self.datos = todos
            nuevos_anios = self.obtener_anios_opciones()
            self.dd_anio.options = nuevos_anios
            if nuevos_anios:
                self.dd_anio.value = nuevos_anios[-1].key

            self.progress_bar.value = 1.0
            self.log(f"🎉 ¡Proceso terminado!")
            self.log(f"📊 Total registros: {len(todos)}")
            self.log(f"➕ Nuevos agregados: {nuevos_contados}")
            self.log(f"🔄 Reemplazos: {reemplazos}")
            self.page.update()
            
            # NO CERRAR AQUÍ. El diálogo permanece abierto.

        except Exception as e:
            self.progress_bar.value = 0
            self.log(f"❌ Error Crítico: {str(e)}")
            self.log("Verifica tu conexión a internet y las credenciales.")
            self.page.update()
        
        # ELIMINADO BLOQUE FINALLY PARA NO CERRAR AUTOMÁTICAMENTE
        # Se cierra solo si el usuario pulsa el botón en el diálogo.

    def parsear_txt(self, texto, estado, sorteo):
        resultados = []
        lineas = texto.strip().splitlines()[1:]  # Saltar encabezado
        for linea in lineas:
            partes = linea.split(",")
            if len(partes) >= 2:
                fecha_str, numeros = partes[0].strip(), partes[1].strip()
                try:
                    fecha = datetime.strptime(fecha_str, "%a %m/%d/%Y")
                except ValueError:
                    continue
                    
                if estado == "FL":
                    digitos = numeros.split("-")
                    if len(digitos) >= 3:
                        numeros = "-".join(digitos[:3])  # Solo los 3 primeros
                            
                fecha_fmt = fecha.strftime("%d/%m/%y")
                # Generar fijos
                nums_parts = numeros.split("-")
                f1 = nums_parts[1] + nums_parts[2]
                f2 = nums_parts[2] + nums_parts[1]
                
                resultados.append({
                    "date": fecha_fmt,
                    "state": estado,
                    "draw": sorteo,
                    "numbers": numeros,
                    "fijos": [f1, f2]
                })
        return resultados

    # --- FIN LÓGICA ACTUALIZACIÓN ---

    def mostrar_paleta_colores(self, e):
        grid = ft.GridView(
            runs_count=6, 
            max_extent=50, 
            run_spacing=10, 
            spacing=10,
            height=300,
        )
        
        dialog = ft.AlertDialog(
            title=ft.Text("Selecciona un Color"),
            content=grid,
            actions=[
                ft.TextButton("Cancelar", on_click=lambda _: self.close_dialog(dialog))
            ]
        )
        
        for color in PALETA_COLORES:
            grid.controls.append(
                ft.Container(
                    width=50, 
                    height=50, 
                    bgcolor=color,
                    border=ft.border.all(2, ft.Colors.BLACK if color != "#FFFFFF" and color != "#FFFF00" and color != "#00FFFF" else ft.Colors.GREY_400),
                    border_radius=10,
                    ink=True,
                    on_click=lambda e, c=color, d=dialog: self.seleccionar_color_manual(c, d)
                )
            )
        
        self.page.show_dialog(dialog)
    
    def seleccionar_color_manual(self, color, dialog):
        self.color_seleccionado_temp = color
        dialog.open = False
        self.page.update()
        self.notificar(f"Color seleccionado: {color}", color)

    def cambiar_visualizacion(self, e):
        if e.control.value:
            self.visualizacion_mode = int(e.control.value)
    
    def actualizar_input_regla(self, e):
        tipo = e.control.value
        if tipo == "completo":
            self.input_digito_container.content = ft.TextField(
                label="Número (00-99)", 
                expand=True, 
                max_length=2, 
                keyboard_type=ft.KeyboardType.NUMBER
            )
        elif tipo == "pareja" or tipo == "digito":
            self.input_digito_container.content = ft.Dropdown(
                label="Dígito", 
                expand=True, 
                options=[ft.dropdown.Option(str(i), str(i)) for i in range(10)]
            )
        else:
            self.input_digito_container.content = ft.Dropdown(
                label="Dígito", 
                expand=True, 
                options=[ft.dropdown.Option(str(i), str(i)) for i in range(10)]
            )
        self.input_digito_container.update()

    def obtener_anios_opciones(self):
        anios = set()
        for sorteo in self.datos:
            try: fecha = datetime.strptime(sorteo["date"], "%d/%m/%y"); anios.add(fecha.year)
            except: continue
        return [ft.dropdown.Option(str(a), str(a)) for a in sorted(anios)]

    def aplicar_filtros(self, e):
        anio_val = self.dd_anio.value
        mes_val = self.dd_mes.value
        state_val = self.dd_state.value
        
        if not anio_val or not state_val:
            self.notificar("Selecciona Año y Lotería", ft.Colors.RED_500)
            return

        self.anio_seleccionado = int(anio_val)
        self.mes_seleccionado = None if mes_val == "Todos" else int(mes_val)
        self.state_seleccionado = state_val

        self.datos_filtrados = []
        for sorteo in self.datos:
            try:
                fecha = datetime.strptime(sorteo["date"], "%d/%m/%y")
                if fecha.year != self.anio_seleccionado: continue
                if self.mes_seleccionado is not None and fecha.month != self.mes_seleccionado: continue
                if sorteo["state"] == self.state_seleccionado: self.datos_filtrados.append(sorteo)
            except: continue
        
        self.renderizar_tabla_anual()
        self.notificar("Filtro aplicado correctamente", ft.Colors.GREEN_500)
    
    def renderizar_tabla_anual(self):
        self.contenedor_vista_tabla.controls.clear()
        
        if self.visualizacion_mode == 0:
            pil_image = self._generar_imagen_tabla_logica()
            if not pil_image: return

            buffer = io.BytesIO()
            pil_image.save(buffer, format="PNG")
            buffer.seek(0)
            img_bytes = buffer.read()
            base64_string = base64.b64encode(img_bytes).decode("utf-8")
            data_uri = f"data:image/png;base64,{base64_string}"
            
            viewer = ft.InteractiveViewer(
                min_scale=0.5,
                max_scale=5.0,
                content=ft.Image(src=data_uri, width=800, fit="contain")
            )
            
            titulo = ft.Text(f"Tabla {self.state_seleccionado} ({self.anio_seleccionado})", size=18, weight=ft.FontWeight.BOLD)
            self.contenedor_vista_tabla.controls.append(titulo)
            self.contenedor_vista_tabla.controls.append(ft.Container(height=10))
            self.contenedor_vista_tabla.controls.append(viewer)
        
        else:
            turnos_state = TURNOS_POR_STATE[self.state_seleccionado]
            meses = [self.mes_seleccionado] if self.mes_seleccionado else list(range(1, 13))
            
            pil_image = self._generar_imagen_horario_completa(turnos_state, self.anio_seleccionado, meses)
            
            if pil_image:
                buffer = io.BytesIO()
                pil_image.save(buffer, format="PNG")
                buffer.seek(0)
                img_bytes = buffer.read()
                base64_string = base64.b64encode(img_bytes).decode("utf-8")
                data_uri = f"data:image/png;base64,{base64_string}"
                
                viewer = ft.InteractiveViewer(
                    min_scale=0.5,
                    max_scale=5.0,
                    content=ft.Image(src=data_uri, width=800, fit="contain")
                )
                
                titulo = ft.Text(f"Tabla {self.state_seleccionado} ({self.anio_seleccionado})", size=18, weight=ft.FontWeight.BOLD)
                self.contenedor_vista_tabla.controls.append(titulo)
                self.contenedor_vista_tabla.controls.append(ft.Container(height=10))
                self.contenedor_vista_tabla.controls.append(viewer)
            
        self.contenedor_vista_tabla.update()

    def _generar_imagen_tabla_logica(self):
        turnos_state = TURNOS_POR_STATE[self.state_seleccionado]
        meses = [self.mes_seleccionado] if self.mes_seleccionado else list(range(1, 13))
        
        datos_map = {}
        for d in range(1, 32):
            datos_map[d] = {}
            for m in meses:
                datos_map[d][m] = {}
                for t in turnos_state: datos_map[d][m][t] = None
        for s in self.datos_filtrados:
            f = datetime.strptime(s["date"], "%d/%m/%y")
            datos_map[f.day][f.month][s["draw"]] = s["fijos"][0]
            
        cell_w_turno = 45 
        cell_w_dia = 30
        cell_h = 35
        margin = 20
        title_height = 30 
        
        total_width = margin + cell_w_dia + (len(meses) * len(turnos_state) * cell_w_turno) + margin
        total_height = margin + (2 * cell_h) + (32 * cell_h) + margin + title_height
        
        try:
            img = Image.new('RGB', (total_width, total_height), color='white')
            draw = ImageDraw.Draw(img)
            font_title = ImageFont.load_default()
            font_header = ImageFont.load_default()
            font_cell = ImageFont.load_default()
        except Exception as e:
            print(e)
            return None

        draw.text((total_width/2, 20), f"{self.state_seleccionado} {self.anio_seleccionado}", fill='black', font=font_title, anchor="mm")
        y = margin + title_height 
        
        draw.rectangle([margin, y, margin+cell_w_dia, y+cell_h], fill='#4472C4', outline='black')
        draw.text((margin+cell_w_dia/2, y+cell_h/2), "Día", fill='white', font=font_header, anchor="mm")
        
        x = margin + cell_w_dia
        for mes in meses:
            w_mes = len(turnos_state) * cell_w_turno
            draw.rectangle([x, y, x+w_mes, y+cell_h], fill='#4472C4', outline='black')
            draw.text((x+w_mes/2, y+cell_h/2), calendar.month_name[mes].upper(), fill='white', font=font_header, anchor="mm")
            x += w_mes
        y += cell_h
        
        draw.rectangle([margin, y, margin+cell_w_dia, y+cell_h], fill='#6085BF', outline='black')
        draw.text((margin+cell_w_dia/2, y+cell_h/2), "T", fill='white', font=font_header, anchor="mm")
        
        x = margin + cell_w_dia
        for mes in meses:
            for t in turnos_state:
                draw.rectangle([x, y, x+cell_w_turno, y+cell_h], fill='#6085BF', outline='black')
                draw.text((x+cell_w_turno/2, y+cell_h/2), t, fill='white', font=font_header, anchor="mm")
                x += cell_w_turno
        y += cell_h
        
        for dia in range(1, 32):
            x = margin
            draw.rectangle([x, y, x+cell_w_dia, y+cell_h], fill='#EEE', outline='black')
            draw.text((x+cell_w_dia/2, y+cell_h/2), str(dia), fill='black', font=font_title, anchor="mm")
            x += cell_w_dia
            
            for m in meses:
                for t in turnos_state:
                    fijo = datos_map[dia][m][t]
                    if fijo:
                        color = self.obtener_color_fijo(fijo) or 'white'
                        txt_col = 'black' if self.es_color_claro(color) else 'white'
                    else:
                        color = 'white'
                        txt_col = 'black'
                    
                    draw.rectangle([x, y, x+cell_w_turno, y+cell_h], fill=color, outline='black')
                    if fijo:
                        draw.text((x+cell_w_turno/2, y+cell_h/2), fijo, fill=txt_col, font=font_cell, anchor="mm")
                    x += cell_w_turno
            y += cell_h

        return img

    def _generar_imagen_horario_completa(self, turnos_state, anio, meses):
        cal = calendar.Calendar(calendar.SUNDAY) 
        
        cell_w = 50 
        cell_h = 30
        margin = 20
        spacer = 20 
        
        col_semana = 40
        col_mes = 50
        col_dia = 40
        
        single_table_width = col_semana + col_mes + (7 * col_dia)
        total_width = (single_table_width * len(turnos_state)) + (spacer * (len(turnos_state) - 1)) + (margin * 2)
        
        total_semanas = 0
        for mes in meses:
            weeks = cal.monthdayscalendar(anio, mes)
            total_semanas += len(weeks)
        
        total_height = margin + cell_h + (total_semanas * cell_h) + margin
        
        fijo_map_global = {}
        for s in self.datos_filtrados:
            f = datetime.strptime(s["date"], "%d/%m/%y")
            key = f.strftime("%Y-%m-%d")
            if key not in fijo_map_global:
                fijo_map_global[key] = {}
            fijo_map_global[key][s["draw"]] = s["fijos"][0]
        
        try:
            img = Image.new('RGB', (int(total_width), int(total_height)), color='white')
            draw = ImageDraw.Draw(img)
            font_header = ImageFont.load_default()
            font_cell = ImageFont.load_default()
            font_bold = ImageFont.load_default()
            font_title = ImageFont.load_default()
        except Exception as e:
            print(e)
            return None

        current_x = margin
        
        for turno in turnos_state:
            draw.text((current_x + single_table_width/2, 10), f"{self.state_seleccionado} {anio} - {turno}", fill='blue', font=font_bold, anchor="mm")
            
            y = margin 
            
            draw.rectangle([current_x, y, current_x+col_semana, y+cell_h], fill='#4472C4', outline='black')
            draw.text((current_x+col_semana/2, y+cell_h/2), "Sem", fill='white', font=font_header, anchor="mm")
            x_temp = current_x + col_semana
            
            draw.rectangle([x_temp, y, x_temp+col_mes, y+cell_h], fill='#4472C4', outline='black')
            draw.text((x_temp+col_mes/2, y+cell_h/2), "Mes", fill='white', font=font_header, anchor="mm")
            x_temp += col_mes
            
            dias_semana = ["D", "L", "M", "M", "J", "V", "S"]
            for d_name in dias_semana:
                draw.rectangle([x_temp, y, x_temp+col_dia, y+cell_h], fill='#4472C4', outline='black')
                draw.text((x_temp+col_dia/2, y+cell_h/2), d_name, fill='white', font=font_header, anchor="mm")
                x_temp += col_dia
            
            y += cell_h
            
            for mes in meses:
                weeks = cal.monthdayscalendar(anio, mes)
                week_num_counter = 1 
                
                for week in weeks:
                    x_temp = current_x
                    
                    draw.rectangle([x_temp, y, x_temp+col_semana, y+cell_h], fill='#EEE', outline='black')
                    draw.text((x_temp+col_semana/2, y+cell_h/2), str(week_num_counter), fill='black', font=font_header, anchor="mm")
                    x_temp += col_semana
                    
                    nombre_mes = ""
                    if 1 in week:
                        nombre_mes = calendar.month_name[mes][:3].upper()
                    
                    draw.rectangle([x_temp, y, x_temp+col_mes, y+cell_h], fill='white', outline='black')
                    if nombre_mes:
                        draw.text((x_temp+col_mes/2, y+cell_h/2), nombre_mes, fill='blue', font=font_bold, anchor="mm")
                    x_temp += col_mes
                    
                    for day in week:
                        color = 'white'
                        txt = ''
                        txt_col = 'black'
                        
                        if day != 0:
                            key = f"{anio}-{mes:02d}-{day:02d}"
                            fijo = fijo_map_global.get(key, {}).get(turno)
                            
                            if fijo:
                                txt = fijo
                                color = self.obtener_color_fijo(fijo) or 'white'
                                txt_col = 'black' if self.es_color_claro(color) else 'white'
                        
                        draw.rectangle([x_temp, y, x_temp+col_dia, y+cell_h], fill=color, outline='black')
                        if txt:
                            draw.text((x_temp+col_dia/2, y+cell_h/2), txt, fill=txt_col, font=font_cell, anchor="mm")
                        x_temp += col_dia
                    
                    y += cell_h
                    week_num_counter += 1
            
            current_x += single_table_width + spacer
                
        return img

    def obtener_color_fijo(self, fijo):
        if not fijo or len(fijo) < 2: return None
        
        regla_completa = next((r for r in self.reglas_colores if r.tipo == "completo" and r.digito == fijo), None)
        if regla_completa:
            return regla_completa.color
        
        regla_pareja = next((r for r in self.reglas_colores if r.tipo == "pareja" and fijo[0] == fijo[1] and r.digito == fijo[0]), None)
        if regla_pareja:
            return regla_pareja.color
        
        regla_digito = next((r for r in self.reglas_colores if r.tipo == "digito" and (fijo[0] == r.digito or fijo[1] == r.digito)), None)
        if regla_digito:
            return regla_digito.color
        
        decena, terminal = fijo[0], fijo[1]
        color_decena = next((r.color for r in self.reglas_colores if r.tipo == "decena" and r.digito == decena), None)
        color_terminal = next((r.color for r in self.reglas_colores if r.tipo == "terminal" and r.digito == terminal), None)
        
        if color_decena and color_terminal: return self.combinar_colores(color_decena, color_terminal)
        return color_decena or color_terminal
    
    def combinar_colores(self, color1, color2):
        r1, g1, b1 = int(color1[1:3], 16), int(color1[3:5], 16), int(color1[5:7], 16)
        r2, g2, b2 = int(color2[1:3], 16), int(color2[3:5], 16), int(color2[5:7], 16)
        return f"#{(r1+r2)//2:02x}{(g1+g2)//2:02x}{(b1+b2)//2:02x}"
    
    def es_color_claro(self, color_str):
        if not color_str: return True
        cl = color_str.lower()
        if cl == 'white': return True
        if cl == 'black': return False
        if color_str.startswith("#"):
            try:
                r, g, b = int(color_str[1:3], 16), int(color_str[3:5], 16), int(color_str[5:7], 16)
                return (0.299 * r + 0.587 * g + 0.114 * b) /255 > 0.5
            except: return True
        return True
    
    def agregar_regla(self, e):
        tipo = self.dd_regla_tipo.value
        input_control = self.input_digito_container.content
        digito_val = input_control.value if isinstance(input_control, ft.TextField) else input_control.value
            
        if not tipo or not digito_val or not hasattr(self, 'color_seleccionado_temp'):
            self.notificar("Completa Tipo, Dígito y Color", ft.Colors.RED_500); return
        
        if tipo == "completo" and len(digito_val) != 2:
             self.notificar("Fijo Completo debe ser 2 dígitos", ft.Colors.RED_500); return
             
        self.reglas_colores.append(ReglaColor(tipo, digito_val, self.color_seleccionado_temp))
        self.guardar_reglas_persistentes()
        self.actualizar_lista_reglas()
        try:
            self.renderizar_tabla_anual()
        except Exception as e:
            print(e)   
        self.notificar("Regla agregada correctamente", ft.Colors.GREEN_500)
    
    def actualizar_lista_reglas(self):
        self.lista_reglas_ui.controls.clear()
        for i, regla in enumerate(self.reglas_colores):
            if regla.tipo == "completo":
                tipo_txt = "Completo"
            elif regla.tipo == "pareja":
                tipo_txt = "Pareja"
            elif regla.tipo == "digito":
                tipo_txt = "Digito"
            elif regla.tipo == "decena":
                tipo_txt = "Decena"
            else:
                tipo_txt = "Terminal"
                
            self.lista_reglas_ui.controls.append(
                ft.Container(
                    padding=5,
                    bgcolor=ft.Colors.GREY_100,
                    border_radius=5,
                    content=ft.Row(controls=[
                        ft.Container(width=15, height=15, bgcolor=regla.color, border_radius=2),
                        ft.Text(f"{tipo_txt} {regla.digito}"),
                        ft.IconButton(icon=ft.Icons.DELETE, icon_color=ft.Colors.RED, on_click=lambda e, idx=i: self.eliminar_regla(idx))
                    ])
                )
            )
        self.lista_reglas_ui.update()
    
    def eliminar_regla(self, idx):
        if 0 <= idx < len(self.reglas_colores):
            del self.reglas_colores[idx]
            self.guardar_reglas_persistentes()
            self.actualizar_lista_reglas()
            try:
                self.renderizar_tabla_anual()
            except Exception as e:
                print(e)    

    # --- EXPORTACIONES ---

    async def exportar_excel(self, e):
        if not self.datos_filtrados: return

        directorio = await ft.FilePicker().get_directory_path(
            dialog_title="Seleccionar carpeta para guardar Excel"
        )
        
        if directorio:
            nombre_archivo = f"tabla_{self.state_seleccionado}_{self.anio_seleccionado}.xlsx"
            ruta_completa = os.path.join(directorio, nombre_archivo)
            
            self._generar_y_guardar_excel(ruta_completa)
            self.notificar(f"Guardado en: {ruta_completa}", ft.Colors.GREEN_500)
        else:
            print("Usuario canceló selección")

    async def exportar_imagen(self, e):
        if not self.datos_filtrados: return

        directorio = await ft.FilePicker().get_directory_path(
            dialog_title="Seleccionar carpeta para guardar Imagen"
        )
        
        if directorio:
            if self.visualizacion_mode == 1:
                turnos_state = TURNOS_POR_STATE[self.state_seleccionado]
                meses = [self.mes_seleccionado] if self.mes_seleccionado else list(range(1,13))
                
                nombre_archivo = f"tabla_{self.state_seleccionado}_completa_{self.anio_seleccionado}.png"
                ruta_completa = os.path.join(directorio, nombre_archivo)
                
                pil_image = self._generar_imagen_horario_completa(turnos_state, self.anio_seleccionado, meses)
                if pil_image:
                    pil_image.save(ruta_completa)
                    self.notificar(f"Imagen guardada: {ruta_completa}", ft.Colors.GREEN_500)
            else:
                nombre_archivo = f"tabla_{self.state_seleccionado}_{self.anio_seleccionado}.png"
                ruta_completa = os.path.join(directorio, nombre_archivo)
                self._generar_y_guardar_imagen(ruta_completa)
                self.notificar(f"Guardado en: {ruta_completa}", ft.Colors.GREEN_500)
        else:
            print("Usuario canceló selección")

    def _generar_y_guardar_excel(self, path):
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Resumen"
            
            if self.visualizacion_mode ==1:
                self._exportar_excel_horario(wb)
            else:
                self._exportar_excel_cubabingo(ws)
                
            wb.save(path)
            self.notificar("Excel guardado exitosamente", ft.Colors.GREEN_500)
        except Exception as ex:
            print(f"Error: {ex}")
            self.notificar(f"Error al guardar Excel: {ex}", ft.Colors.RED_500)

    def _exportar_excel_cubabingo(self, ws):
        turnos_state = TURNOS_POR_STATE[self.state_seleccionado]
        meses = [self.mes_seleccionado] if self.mes_seleccionado else list(range(1, 13))
        
        datos_map = {}
        for d in range(1, 32):
            datos_map[d] = {}
            for m in meses:
                datos_map[d][m] = {}
                for t in turnos_state: datos_map[d][m][t] = None
        for s in self.datos_filtrados:
            f = datetime.strptime(s["date"], "%d/%m/%y")
            datos_map[f.day][f.month][s["draw"]] = s["fijos"][0]
            
        col = 1
        ws.cell(row=1, column=col, value="Día").alignment = Alignment(horizontal="center", vertical="center")
        col += 1
        
        for mes in meses:
            start_col = col
            ws.cell(row=1, column=start_col, value=calendar.month_name[mes].upper()).alignment = Alignment(horizontal="center", vertical="center")
            end_col = col + len(turnos_state) - 1
            if start_col < end_col: ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            col = end_col + 1
            
        col = 1
        ws.cell(row=2, column=col, value="Turno").alignment = Alignment(horizontal="center")
        col += 1
        for mes in meses:
            for t in turnos_state:
                ws.cell(row=2, column=col, value=t).alignment = Alignment(horizontal="center")
                col += 1
                
        fill_header = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        font_header = Font(bold=True, color="FFFFFF")
        for r in [1, 2]:
            for c in range(1, col):
                ws.cell(row=r, column=c).fill = fill_header
                ws.cell(row=r, column=c).font = font_header

        row_idx = 3
        for dia in range(1, 32):
            ws.cell(row=row_idx, column=1, value=str(dia)).alignment = Alignment(horizontal="center")
            col_idx = 2
            for m in meses:
                for t in turnos_state:
                    fijo = datos_map[dia][m][t]
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if fijo:
                        cell.value = fijo
                        color = self.obtener_color_fijo(fijo)
                        if color:
                            c_code = color[1:] if color.startswith("#") else color
                            try: cell.fill = PatternFill(start_color=c_code, end_color=c_code, fill_type="solid")
                            except: pass
                            if not self.es_color_claro(color): cell.font = Font(color="FFFFFF", bold=True)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    col_idx += 1
            row_idx += 1
            
    def _exportar_excel_horario(self, wb):
        
        turnos_state = TURNOS_POR_STATE[self.state_seleccionado]
        meses = [self.mes_seleccionado] if self.mes_seleccionado else list(range(1, 13))
        
        fijo_map_global = {}
        for s in self.datos_filtrados:
            f = datetime.strptime(s["date"], "%d/%m/%y")
            key = f"{f.year}-{f.month:02d}-{f.day:02d}"
            if key not in fijo_map_global:
                fijo_map_global[key] = {}
            fijo_map_global[key][s["draw"]] = s["fijos"][0]
            
        cal = calendar.Calendar(calendar.SUNDAY)
        
        cols_per_turn = 9
        turnos_offsets = {}
        current_col = 1
        for i, t in enumerate(turnos_state):
            turnos_offsets[t] = current_col
            current_col += cols_per_turn
        
        for t in turnos_state:
            start = turnos_offsets[t]
            end = start + cols_per_turn - 1
            ws_cell = wb.active.cell(row=1, column=start, value=f"Turno {t}")
            ws_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws_cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            ws_cell.font = Font(bold=True, color="FFFFFF")
            if start < end:
                wb.active.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        
        headers = ["Semana", "Mes", "Dom", "Lun", "Mar", "Mié", "Jue", "Vie", "Sáb"]
        for t in turnos_state:
            base_col = turnos_offsets[t]
            for c_idx, h in enumerate(headers):
                cell = wb.active.cell(row=2, column=base_col + c_idx, value=h)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center")
        
        row_idx = 3
        for mes in meses:
            weeks = cal.monthdayscalendar(self.anio_seleccionado, mes)
            week_num = 1
            for week in weeks:
                for t in turnos_state:
                    base_col = turnos_offsets[t]
                    
                    wb.active.cell(row=row_idx, column=base_col, value=week_num).alignment = Alignment(horizontal="center")
                    
                    if 1 in week:
                        cell_m = wb.active.cell(row=row_idx, column=base_col + 1, value=calendar.month_name[mes])
                        cell_m.alignment = Alignment(horizontal="center")
                        cell_m.font = Font(bold=True, color="0000FF")
                    
                    for day_idx, day in enumerate(week):
                        col_idx = base_col + 2 + day_idx
                        if day != 0:
                            key = f"{self.anio_seleccionado}-{mes:02d}-{day:02d}"
                            fijo = fijo_map_global.get(key, {}).get(t)
                            
                            if fijo:
                                cell = wb.active.cell(row=row_idx, column=col_idx, value=fijo)
                                color = self.obtener_color_fijo(fijo)
                                if color:
                                    c_code = color[1:] if color.startswith("#") else color
                                    try: cell.fill = PatternFill(start_color=c_code, end_color=c_code, fill_type="solid")
                                    except: pass
                                    if not self.es_color_claro(color): cell.font = Font(color="FFFFFF", bold=True)
                                cell.alignment = Alignment(horizontal="center")
                
                row_idx +=1
                week_num += 1

    def _generar_y_guardar_imagen(self, path):
        try:
            pil_image = self._generar_imagen_tabla_logica()
            if pil_image:
                pil_image.save(path)
                self.notificar("Imagen guardada exitosamente", ft.Colors.GREEN_500)
            else:
                self.notificar("Error generando la imagen", ft.Colors.RED_500)
        except Exception as ex:
            print(f"Error: {ex}")
            self.notificar(f"Error al guardar Imagen: {ex}", ft.Colors.RED_500)

    def notificar(self, mensaje, color):
        self.page.show_dialog(ft.SnackBar(mensaje))
    
    def mostrar_ayuda(self, e):
        d = ft.AlertDialog(
            title=ft.Text("Ayuda"), 
            content=ft.Text("Selecciona el tipo de visualización:\n- Cubabingo: Vista anual tradicional.\n- Por Horario: Tablas de turnos (M, E, N) una al lado de la otra.\n\nNuevas Reglas:\n- Pareja: Colorea solo dobles (22, 33, etc).\n- Digito: Colorea cualquier aparición del dígito.\n\nUsa dos dedos para hacer Zoom en la tabla.\nPara exportar, selecciona la carpeta Descargas en el diálogo."), 
            actions=[ft.TextButton("Cerrar", on_click=lambda _: self.close_dialog(d))]
        )
        self.page.show_dialog(d)

    def close_dialog(self, d):
        d.open = False
        self.progress_bar.value = 0 # Resetear barra de progreso al cerrar manualmente
        self.page.update()

def main(page: ft.Page):
    LoteriaApp(page)

ft.run(main)
